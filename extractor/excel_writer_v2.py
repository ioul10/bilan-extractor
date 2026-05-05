# ============================================================
# excel_writer_v2.py
# Remplit le MODELE.xlsx avec les données extraites et validées
#
# Principe :
#   - Copie le MODELE.xlsx (template avec formules + styles)
#   - Remplit les cellules de saisie avec les valeurs extraites
#   - Les formules recalculent automatiquement les totaux
#   - Ajoute les infos d'identification (société, date...)
#   - Ajoute un onglet Rapport de validation
# ============================================================

import io
import copy
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

# ── MAPPING LIGNE → COLONNES DE SAISIE ──────────────────────
# Pour chaque section, quelles cellules sont en saisie manuelle
# (pas des formules) et quelles colonnes remplir

# Actif : B=Brut, C=Amort — D et E sont calculés par formule
ACTIF_INPUT_COLS = {
    "B": "brut",   # Colonne Brut → saisie
    "C": "amort",  # Colonne Amort → saisie
    # D = Net N   → calculé par formule (B-C)
    # E = Net N-1 → saisie directe
    "E": "net_n1",
}

# Lignes Actif en saisie (pas des totaux/formules)
ACTIF_INPUT_ROWS = [
    7, 8, 9,           # [A] Non valeurs
    11, 12, 13, 14,    # [B] Incorporelles
    16, 17, 18, 19, 20, 21, 22,  # [C] Corporelles
    24, 25, 26, 27,    # [D] Financières
    29, 30,            # [E] Écarts conversion
    33, 34, 35, 36, 37,  # [F] Stocks
    39, 40, 41, 42, 43, 44, 45,  # [G] Créances
    46,                # [H] TVP
    47,                # [I] Écarts circulants
    50, 51, 52,        # Trésorerie
]

# Passif : B=Exercice N, C=Exercice N-1
PASSIF_INPUT_ROWS = [
    7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17,  # Capitaux propres
    20, 21,            # Cap propres assimilés
    23, 24,            # Dettes financement
    26, 27,            # Provisions durables
    29, 30,            # Écarts conversion
    33, 34, 35, 36, 37, 38, 39, 40,  # Passif circulant
    41,                # Autres provisions
    42,                # Écarts circulants
    45, 46, 47,        # Trésorerie passif
]

# CPC : C=Propres, D=Précédents, F=Exercice N-1
# E = Total N est calculé automatiquement (C+D)
CPC_INPUT_ROWS = [
    7, 8, 9, 10, 11, 12, 13, 14,   # Produits exploitation
    17, 18, 19, 20, 21, 22, 23,    # Charges exploitation
    27, 28, 29, 30,                 # Produits financiers
    33, 34, 35, 36,                 # Charges financières
    41, 42, 43, 44, 45,            # Produits non courants
    48, 49, 50, 51,                # Charges non courants
    55,                            # Impôts
]


# ── COULEURS DU RAPPORT ──────────────────────────────────────
COLOR_OK      = "C6EFCE"   # Vert clair
COLOR_ERROR   = "FFCCCC"   # Rouge clair
COLOR_WARNING = "FFEB9C"   # Orange clair
COLOR_HEADER  = "1F3864"   # Bleu foncé
COLOR_WHITE   = "FFFFFF"


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ── EXTRACTION DES VALEURS PAR LIGNE ────────────────────────

def _get_value(data: dict, ligne: int, col_key: str) -> float | None:
    """
    Récupère une valeur dans le data_dict.
    col_key : "B", "C", "D", "E", "F"
    """
    if ligne not in data:
        return None
    return data[ligne].get(col_key)


def _map_actif_to_cells(actif_data: dict) -> dict:
    """
    Construit le mapping cellule → valeur pour l'Actif.
    Retourne : { "B7": 501032290.18, "C7": 282447118.19, ... }
    """
    cells = {}
    for ligne in ACTIF_INPUT_ROWS:
        # Brut
        b = _get_value(actif_data, ligne, "B")
        if b is not None:
            cells[f"B{ligne}"] = b

        # Amort
        c = _get_value(actif_data, ligne, "C")
        if c is not None:
            cells[f"C{ligne}"] = c

        # Net N-1
        e = _get_value(actif_data, ligne, "E")
        if e is not None:
            cells[f"E{ligne}"] = e

    return cells


def _map_passif_to_cells(passif_data: dict) -> dict:
    """Mapping cellule → valeur pour le Passif."""
    cells = {}
    for ligne in PASSIF_INPUT_ROWS:
        b = _get_value(passif_data, ligne, "B")
        if b is not None:
            cells[f"B{ligne}"] = b

        c = _get_value(passif_data, ligne, "C")
        if c is not None:
            cells[f"C{ligne}"] = c

    return cells


def _map_cpc_to_cells(cpc_data: dict) -> dict:
    """Mapping cellule → valeur pour le CPC."""
    cells = {}
    for ligne in CPC_INPUT_ROWS:
        # Propres à l'exercice
        c = _get_value(cpc_data, ligne, "C")
        if c is not None:
            cells[f"C{ligne}"] = c

        # Exercices précédents
        d = _get_value(cpc_data, ligne, "D")
        if d is not None:
            cells[f"D{ligne}"] = d

        # Totaux exercice N-1
        f = _get_value(cpc_data, ligne, "F")
        if f is not None:
            cells[f"F{ligne}"] = f

    return cells


# ── ONGLET RAPPORT ───────────────────────────────────────────

def _add_rapport_sheet(wb, rapport: dict):
    """Ajoute un onglet Rapport de validation au classeur."""

    ws = wb.create_sheet("Rapport Validation")
    ws.sheet_properties.tabColor = "FF0000" if rapport["score_global"] < 80 else "00B050"

    # En-tête
    ws["A1"] = "RAPPORT DE VALIDATION BILAN"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Score global : {rapport['score_global']}%"
    ws["A2"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:F2")

    # Équilibre
    eq = rapport["equilibre"]
    row = 4
    ws.cell(row, 1, "ÉQUILIBRE FONDAMENTAL").font = Font(bold=True)
    row += 1
    ws.cell(row, 1, eq["message"])
    if eq.get("total_actif"):
        ws.cell(row, 2, eq["total_actif"]).number_format = '#,##0.00'
        ws.cell(row, 3, "=")
        ws.cell(row, 4, eq["total_passif"]).number_format = '#,##0.00'
    ok_color = COLOR_OK if eq.get("ok") else COLOR_ERROR
    for col in range(1, 5):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=ok_color)

    # Sections
    row += 2
    for section_key in ["actif", "passif", "cpc"]:
        rep = rapport[section_key]

        # En-tête section
        ws.cell(row, 1, f"{rep['section'].upper()} — Score : {rep['score']}%")
        ws.cell(row, 1).font = Font(bold=True, color=COLOR_WHITE)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLOR_HEADER)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        # En-tête colonnes
        for col, header in enumerate(
            ["Ligne", "Libellé", "Extrait", "Calculé", "Écart", "Status"], 1
        ):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
        row += 1

        # Lignes
        for l in rep["lignes"]:
            status_map = {"ok": "✅", "error": "❌", "missing": "⚠️"}
            icon = status_map.get(l["status"], "")
            color_map = {
                "ok":      COLOR_OK,
                "error":   COLOR_ERROR,
                "missing": COLOR_WARNING,
            }
            bg = color_map.get(l["status"], COLOR_WHITE)

            ws.cell(row, 1, l["ligne"])
            ws.cell(row, 2, l["label"])
            ws.cell(row, 3, l["extrait"] or "")
            ws.cell(row, 4, l["calculé"] or "")
            ws.cell(row, 5, l["écart"] or "")
            ws.cell(row, 6, icon)

            for col in range(1, 7):
                c = ws.cell(row, col)
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = _thin_border()
                c.font = Font(size=9)
                if col in [3, 4, 5] and ws.cell(row, col).value:
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")

            row += 1

        row += 1

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 8


# ── FONCTION PRINCIPALE ──────────────────────────────────────

def fill_model(
    model_path: str,
    actif_data: dict,
    passif_data: dict,
    cpc_data: dict,
    rapport: dict,
    identification: dict = None,
) -> bytes:
    """
    Remplit le MODELE.xlsx avec les données extraites.

    model_path     : chemin vers MODELE.xlsx (template)
    actif_data     : { ligne: {col: valeur} }
    passif_data    : idem
    cpc_data       : idem
    rapport        : sortie de validate_all()
    identification : { "raison_sociale": ..., "date_bilan": ..., ... }

    Retourne les bytes du fichier Excel rempli.
    """
    wb = load_workbook(model_path)

    # ── Identification ───────────────────────────────────────
    if identification:
        ws_id = wb["1 - Identification"]
        id_map = {
            "Raison sociale":      identification.get("raison_sociale", ""),
            "Identifiant fiscal":  identification.get("identifiant_fiscal", ""),
            "Taxe professionnelle": identification.get("taxe_pro", ""),
            "Adresse":             identification.get("adresse", ""),
            "Date de bilan":       identification.get("date_bilan", ""),
            "Format PDF":          identification.get("format_pdf", ""),
        }
        for row in ws_id.iter_rows():
            for cell in row:
                if cell.value and str(cell.value) in id_map:
                    # La valeur est dans la cellule suivante
                    next_cell = ws_id.cell(cell.row, cell.column + 1)
                    next_cell.value = id_map[str(cell.value)]

        # Date de bilan dans Actif et Passif
        date = identification.get("date_bilan", "")
        if date:
            wb["2 - Bilan Actif"]["A3"] = f"Date de bilan : {date}"
            wb["3 - Bilan Passif"]["A3"] = f"Date de bilan : {date}"
            wb["4 - CPC"]["A3"] = f"Date de bilan : {date}"

        # IF dans Actif/Passif/CPC
        if_val = identification.get("identifiant_fiscal", "")
        if if_val:
            wb["2 - Bilan Actif"]["E2"]  = f"IF : {if_val}"
            wb["3 - Bilan Passif"]["C2"] = f"IF : {if_val}"
            wb["4 - CPC"]["F2"]          = f"IF : {if_val}"

    # ── Actif ────────────────────────────────────────────────
    ws_actif = wb["2 - Bilan Actif"]
    actif_cells = _map_actif_to_cells(actif_data)
    for cell_ref, value in actif_cells.items():
        ws_actif[cell_ref] = value
        ws_actif[cell_ref].number_format = '#,##0.00'

    # ── Passif ───────────────────────────────────────────────
    ws_passif = wb["3 - Bilan Passif"]
    passif_cells = _map_passif_to_cells(passif_data)
    for cell_ref, value in passif_cells.items():
        ws_passif[cell_ref] = value
        ws_passif[cell_ref].number_format = '#,##0.00'

    # ── CPC ──────────────────────────────────────────────────
    ws_cpc = wb["4 - CPC"]
    cpc_cells = _map_cpc_to_cells(cpc_data)
    for cell_ref, value in cpc_cells.items():
        ws_cpc[cell_ref] = value
        ws_cpc[cell_ref].number_format = '#,##0.00'

    # ── Rapport de validation ────────────────────────────────
    _add_rapport_sheet(wb, rapport)

    # ── Sauvegarder en mémoire ───────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
