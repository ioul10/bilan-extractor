import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Couleurs ──────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F3864"   # bleu foncé
COLOR_HEADER_FG   = "FFFFFF"   # blanc
COLOR_SECTION_BG  = "D6E4F0"   # bleu clair
COLOR_TOTAL_BG    = "BDD7EE"   # bleu moyen
COLOR_ALT_ROW     = "F2F7FB"   # gris très clair (alternance)
COLOR_EMPTY_SHEET = "F5F5F5"   # fond onglet vide

# ── Mots-clés pour détecter les lignes importantes ────────────────
TOTAL_KEYWORDS = ["total", "total i", "total ii", "total iii",
                  "total général", "total general", "résultat"]
SECTION_KEYWORDS = ["immobilisations", "stocks", "créances", "crédits",
                    "capitaux propres", "dettes", "trésorerie",
                    "produits", "charges", "exploitation", "financier"]


def _is_total_row(row: list) -> bool:
    label = (row[0] or "").lower()
    return any(kw in label for kw in TOTAL_KEYWORDS)


def _is_section_row(row: list) -> bool:
    label = (row[0] or "").lower()
    # Ligne section = texte en majuscules ou contient mot-clé
    return any(kw in label for kw in SECTION_KEYWORDS)


def _is_numeric(val: str) -> bool:
    if not val:
        return False
    cleaned = val.replace(" ", "").replace(",", ".").replace("-", "").strip()
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _parse_number(val: str):
    """Convertit une chaîne marocaine en float : '1 234 567,89' → 1234567.89"""
    if not val or not val.strip():
        return None
    s = val.strip().replace("\xa0", "").replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1]
    if s.startswith("-"):
        neg, s = True, s[1:]
    s = s.replace(",", ".")
    try:
        result = float(s)
        return -result if neg else result
    except ValueError:
        return None


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_identification_sheet(wb: Workbook):
    """Crée l'onglet Identification — vide, prêt pour le dictionnaire."""
    ws = wb.create_sheet("Identification", 0)
    ws.sheet_properties.tabColor = "4472C4"

    # Titre
    ws["B2"] = "IDENTIFICATION"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=COLOR_HEADER_BG)

    # Message placeholder
    ws["B4"] = "Cet onglet sera complété lors de l'étape de normalisation (dictionnaire des clés)."
    ws["B4"].font = Font(name="Calibri", size=10, italic=True, color="888888")

    # Fond
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLOR_EMPTY_SHEET)

    ws.column_dimensions["B"].width = 70


def _write_data_sheet(wb: Workbook, sheet_name: str, rows: list[list],
                      tab_color: str = "2E75B6"):
    """
    Écrit une section (Actif/Passif/CPC) dans un onglet Excel.
    - Première ligne détectée = en-tête de colonnes
    - Valeurs numériques → format nombre
    - Lignes TOTAL → mise en évidence
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    if not rows:
        ws["A1"] = "Aucune donnée extraite."
        return

    # Détecter le nombre de colonnes max
    max_cols = max(len(row) for row in rows)

    # ── En-tête (première ligne non vide) ────────────────────────
    header_written = False
    data_start = 0

    for i, row in enumerate(rows):
        if any(cell.strip() for cell in row if cell):
            # Première ligne valide = en-tête
            ws.append(row + [""] * (max_cols - len(row)))
            for col_idx in range(1, max_cols + 1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color=COLOR_HEADER_FG)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
                cell.border = _thin_border()
            ws.row_dimensions[ws.max_row].height = 35
            header_written = True
            data_start = i + 1
            break

    # ── Données ──────────────────────────────────────────────────
    for row_idx, row in enumerate(rows[data_start:], start=1):
        padded = row + [""] * (max_cols - len(row))
        ws.append(padded)
        excel_row = ws.max_row
        is_total = _is_total_row(padded)
        is_section = _is_section_row(padded)
        is_alt = row_idx % 2 == 0

        for col_idx, val in enumerate(padded, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=9)

            # ── Valeur numérique ──
            if col_idx > 1 and _is_numeric(val):
                num = _parse_number(val)
                if num is not None:
                    cell.value = num
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = val
            else:
                cell.value = val
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left",
                                               indent=1 if not is_total else 0)

            # ── Style selon type de ligne ──
            if is_total:
                cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
                cell.font = Font(name="Calibri", size=9, bold=True)
            elif is_section:
                cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
                cell.font = Font(name="Calibri", size=9, bold=True)
            elif is_alt:
                cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

        # Ligne séparateur de page
        if padded[0] == "--- PAGE SUIVANTE ---":
            for col_idx in range(1, max_cols + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor="FFE699")
                cell.font = Font(italic=True, size=8, color="888888")

    # ── Largeur des colonnes ──────────────────────────────────────
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 45
        else:
            ws.column_dimensions[col_letter].width = 20

    # Figer la première ligne
    ws.freeze_panes = "A2"


def build_excel(sections: dict) -> bytes:
    """
    Construit le fichier Excel en mémoire.

    sections = {
        "Actif":  [ [row], [row], ... ],
        "Passif": [ [row], ... ],
        "CPC":    [ [row], ... ],
    }

    Retourne les bytes du fichier .xlsx
    """
    wb = Workbook()
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Onglet 1 : Identification (vide)
    _write_identification_sheet(wb)

    # Couleurs par section
    tab_colors = {
        "Actif":  "2E75B6",   # bleu
        "Passif": "375623",   # vert foncé
        "CPC":    "7030A0",   # violet
    }

    # Onglets 2, 3, 4
    for name, rows in sections.items():
        color = tab_colors.get(name, "4472C4")
        _write_data_sheet(wb, name, rows, tab_color=color)

    # Sauvegarder en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
